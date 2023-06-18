from datetime import datetime
from openpyxl import load_workbook
import glob
import re
import pandas as pd
from time import sleep
import requests
from urllib.parse import urlparse, parse_qs

pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)

requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = (         # pyright: ignore
    "ALL:@SECLEVEL=1"  
)

col_names = ["Sex", "Type", "Age", "Date", "Value"]
df = pd.DataFrame(columns=col_names)


def find_data_files():
    file_list = glob.glob("./../raw_data/*.xlsx")

    matching_files = []
    for file_name in file_list:
        if re.search("UPRAWNIENIA", file_name):
            matching_files.append(file_name)

    return matching_files


find_data_files()


def get_data(file_name, df):
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    year_col = ""
    value_col = ""
    year = 0

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if (cell.row == 1) and (
                re.search("data stanu", str(cell.value), re.IGNORECASE)
                or re.search("data_stanu", str(cell.value), re.IGNORECASE)
            ):
                year_col = cell.column_letter
            if (cell.row == 1) and re.search("liczba", str(cell.value), re.IGNORECASE):
                value_col = cell.column_letter

    if len(year_col) == 0:
        matches = re.search(r"_(\d{4})\.", file_name)
        if matches:
            year = matches.group(1)

    for row in ws.iter_rows(min_row=2):
        sex = ""
        type = ""
        age = ""
        value = 0
        for cell in row:
            if cell.column == 1:
                sex_string = cell.value
                if sex_string == "MĘŻCZYZNA" or sex_string == "MEZCZYZNA":
                    sex = "MĘŻCZYŹNI"
                elif sex_string == "KOBIETA":
                    sex = "KOBIETY"
                else:
                    sex = cell.value
            if cell.column == 2:
                type = cell.value
            if cell.column == 3:
                age_str = cell.value
                age = age_str.lower()
            if cell.column_letter == year_col:
                date_string = str(cell.value).split()[0]
                try:
                    date_obj = datetime.strptime(date_string, "%Y-%m-%d")
                    year = date_obj.year
                except:
                    pass
            if cell.column_letter == value_col:
                value = cell.value
        year = int(year)
        value = int(value)
        new_row = [sex, type, age, year, value]
        df.loc[len(df)] = new_row


def get_data_year_2022():
    col_names_extended = col_names
    col_names_extended.append("Region")


    tmp_df = pd.DataFrame(columns=col_names_extended)

    for i in range(12):
        data_url = f"https://api.cepik.gov.pl/uprawnienia?limit=500&filter[data-statystyki]=2022-{str(i+1).zfill(2)}"
        total_number_of_pages_link = ""

        while True:
            try: 
                total_number_of_pages_link = requests.get(data_url).json()["links"]["last"]
                break;
            except:
                sleep(10) 

        parsed_link = urlparse(total_number_of_pages_link)
        link_params = parse_qs(parsed_link.query)

        max_pages = int(link_params["page"][0])
        j = 0

        while j < max_pages:
            data_url_paginated = data_url + f"&page={j+1}" 
            try:
                print(f"trying to get data from page {j+1} for month {i+1}")
                parsed_data = requests.get(data_url_paginated).json()
                for element in parsed_data["data"]:
                    element = element["attributes"]
                    sex = ""
                    if element["plec"] == "K":
                        sex = "KOBIETY"
                    else:
                        sex = "MĘŻCZYŹNI"
                    age = ""
                    age_parsed = int(element["wiek"])
                    if age_parsed >= 18 and age_parsed <= 24:
                        age = "18-24"
                    elif age_parsed >= 25 and age_parsed <= 34:
                        age = "25-34"
                    elif age_parsed >= 35 and age_parsed <=44:
                        age = "35-44"
                    elif age_parsed >= 45 and age_parsed <=54:
                        age = "45-54"
                    elif age_parsed >= 55 and age_parsed <= 64:
                        age = "55-64"
                    else:
                        age = "od 65"

                    general_data_row = [
                        sex,
                        element["kod-uprawnienia"],
                        age,
                        element["data-statystyki"],
                        element["ilosc"],
                        element['wojewodztwo-nazwa'],
                    ]
                    tmp_df.loc[len(tmp_df)] = general_data_row
                j+=1
            except:
                print("program crashed retrying in 10 seconds")
                sleep(10)
    tmp_df.to_csv("./../data/uprawnienia_2022.csv", index=False)


def process_data(df):
    work_df = pd.read_csv("./../data/uprawnienia_2022.csv")
    general_df = work_df.groupby(["Sex", "Type", "Age"])["Value"].sum().reset_index()
    general_df.insert(3, "Date", 2022)
    return pd.concat([df, general_df]).reset_index()

# get_data_year_2022()

for file_name in find_data_files():
    get_data(file_name, df)
df = process_data(df)
df = df.sort_values(by='Date')

print(df)
df.to_csv('./../data/uprawnienia.csv', index=False)
